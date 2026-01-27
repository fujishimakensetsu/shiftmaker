# -*- coding: utf-8 -*-
"""
Excel出力サービス
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from models import get_locations, get_staff
from .calendar_service import get_calendar_data


def create_excel_shift(year, month, shift_data, month_exceptions):
    """Excel形式のシフト表を作成"""
    locations = get_locations()
    staff_list = get_staff()
    staff_dict = {s['id']: s['name'] for s in staff_list}
    num_locations = len(locations)

    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}年{month}月シフト表"

    # スタイル定義
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    saturday_fill = PatternFill(start_color='DEEAF6', end_color='DEEAF6', fill_type='solid')
    sunday_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
    holiday_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
    closed_day_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
    loc_name_fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')

    saturday_font = Font(color='0000FF', bold=True)
    sunday_font = Font(color='FF0000', bold=True)

    cal_data = get_calendar_data(year, month, month_exceptions)
    weekday_headers = ['日', '月', '火', '水', '木', '金', '土']

    # カレンダーデータを週単位でグループ化
    weeks = []
    current_week = [None] * 7
    for day_info in cal_data:
        weekday_jp = (day_info['weekday'] + 1) % 7
        current_week[weekday_jp] = day_info
        if weekday_jp == 6:
            weeks.append(current_week)
            current_week = [None] * 7
    if any(d is not None for d in current_week):
        weeks.append(current_week)

    # 列幅設定
    ws.column_dimensions['A'].width = 6
    for col in range(2, 9):
        ws.column_dimensions[get_column_letter(col)].width = 12

    # ヘッダー行
    ws.cell(row=1, column=1, value="").border = thin_border
    for col, day_name in enumerate(weekday_headers, start=2):
        cell = ws.cell(row=1, column=col, value=day_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if col == 2:
            cell.fill = PatternFill(start_color='dc3545', end_color='dc3545', fill_type='solid')
        elif col == 8:
            cell.fill = PatternFill(start_color='0d6efd', end_color='0d6efd', fill_type='solid')
    ws.row_dimensions[1].height = 20

    current_row = 2

    for week_idx, week in enumerate(weeks):
        # 全拠点が定休日かチェック
        all_closed_days = []
        for day_info in week:
            if day_info:
                all_closed = all(loc_info.get('is_closed_day', False) for loc_info in day_info['locations'])
                all_closed_days.append(all_closed)
            else:
                all_closed_days.append(False)

        # 日付行
        date_row = current_row
        month_cell = ws.cell(row=date_row, column=1)
        if week_idx == 0:
            month_cell.value = f"{month}月"
            month_cell.font = Font(size=12, bold=True)
        else:
            month_cell.value = ""
        month_cell.alignment = Alignment(horizontal='center', vertical='center')
        month_cell.border = thin_border

        for col_idx, day_info in enumerate(week):
            col = col_idx + 2
            cell = ws.cell(row=date_row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            if day_info:
                day = day_info['day']
                weekday_jp = (day_info['weekday'] + 1) % 7
                is_holiday = day_info['is_holiday']
                holiday_name = day_info['holiday_name']

                if is_holiday and holiday_name:
                    cell.value = f"{day}\n{holiday_name}"
                else:
                    cell.value = day

                if is_holiday or weekday_jp == 0:
                    cell.font = sunday_font
                    cell.fill = sunday_fill if not is_holiday else holiday_fill
                elif weekday_jp == 6:
                    cell.font = saturday_font
                    cell.fill = saturday_fill
                else:
                    cell.font = Font(bold=True)

        ws.row_dimensions[date_row].height = 30

        # 各拠点行
        for loc_idx, loc in enumerate(locations):
            loc_row = date_row + 1 + loc_idx

            loc_name_cell = ws.cell(row=loc_row, column=1, value=loc['name'])
            loc_name_cell.font = Font(size=9, bold=True)
            loc_name_cell.alignment = Alignment(horizontal='center', vertical='center')
            loc_name_cell.border = thin_border
            loc_name_cell.fill = loc_name_fill

            for col_idx, day_info in enumerate(week):
                col = col_idx + 2
                cell = ws.cell(row=loc_row, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                if day_info:
                    date_str = day_info['date']
                    loc_info = next((l for l in day_info['locations'] if l['id'] == loc['id']), None)

                    if all_closed_days[col_idx]:
                        if loc_idx == 0:
                            if num_locations > 1:
                                ws.merge_cells(start_row=loc_row, start_column=col,
                                             end_row=loc_row + num_locations - 1, end_column=col)
                            cell.value = "定休日"
                            cell.font = Font(size=11, bold=True)
                            cell.fill = closed_day_fill
                    elif loc_info and loc_info.get('is_closed_day', False):
                        cell.value = "休"
                        cell.fill = closed_day_fill
                    elif loc_info and not loc_info['is_working']:
                        cell.value = ""
                    else:
                        shift_for_day = shift_data.get(date_str, {})
                        assigned_ids = shift_for_day.get(str(loc['id']), [])
                        if assigned_ids:
                            names = [staff_dict.get(int(sid) if isinstance(sid, str) else sid, '?')
                                   for sid in assigned_ids if sid]
                            cell.value = "\n".join(names) if names else "-"
                            cell.font = Font(size=9, bold=True)
                        else:
                            cell.value = "-"

            ws.row_dimensions[loc_row].height = 26

        current_row = date_row + 1 + num_locations

    return wb
