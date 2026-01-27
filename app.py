# -*- coding: utf-8 -*-
"""
建築会社 展示場シフト表自動作成アプリ
Flask + Bootstrap 5 + openpyxl
Cloud Run対応版（Firestore + パスワード認証）
"""

import os
import json
import calendar
from datetime import datetime, date, timedelta
from io import BytesIO
from copy import deepcopy
from pathlib import Path
from functools import wraps

from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for, session, make_response
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
import jpholiday
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# PDF出力用
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# Firestore（ローカル開発時は無効化）
FIRESTORE_AVAILABLE = False
try:
    if os.environ.get('GOOGLE_CLOUD_PROJECT'):
        from google.cloud import firestore
        FIRESTORE_AVAILABLE = True
except ImportError:
    pass

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production')

# セッション設定
app.config['SESSION_COOKIE_SECURE'] = False  # HTTPSでない場合はFalse
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['REMEMBER_COOKIE_DURATION'] = timedelta(days=7)

# =============================================================================
# 認証設定（Flask-Login）
# =============================================================================

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'ログインしてください'
login_manager.session_protection = 'basic'

# 環境変数からパスワードを取得（デフォルトは開発用）
APP_PASSWORD = os.environ.get('APP_PASSWORD', 'shift2026')


class User(UserMixin):
    """シンプルなユーザークラス（単一ユーザー認証用）"""
    def __init__(self, user_id):
        self.id = user_id

    def get_id(self):
        return str(self.id)


@login_manager.user_loader
def load_user(user_id):
    if user_id == 'admin':
        return User('admin')
    return None


@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))

    error = None
    if request.method == 'POST':
        password = request.form.get('password', '')
        if password == APP_PASSWORD:
            user = User('admin')
            login_user(user, remember=True)
            next_page = request.args.get('next') or '/'
            return redirect(next_page)
        else:
            error = 'パスワードが正しくありません'

    return render_template('login.html', error=error)


@app.route('/')
@login_required
def index():
    return render_template('index.html',
                         locations=get_locations(),
                         staff=get_staff())


@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

# =============================================================================
# データ管理（Firestore優先、ローカルJSONフォールバック）
# =============================================================================

DATA_DIR = Path(__file__).parent / 'data'
DATA_FILE = DATA_DIR / 'settings.json'

DEFAULT_DATA = {
    "locations": [
        {
            "id": 1,
            "name": "FIP",
            "working_days": [0, 1, 2, 3, 4, 5, 6],
            "closed_days": [2, 3],  # 定休日（水曜、木曜）
            "work_on_holidays": True,
            "min_staff": 1,
            "max_staff": 2,
            "part_time_priority": False,
            "flexible_staffing": False
        },
        {
            "id": 2,
            "name": "大宮",
            "working_days": [5, 6],
            "closed_days": [2, 3],  # 定休日（水曜、木曜）
            "work_on_holidays": True,
            "min_staff": 1,
            "max_staff": 2,
            "part_time_priority": True,
            "flexible_staffing": True
        },
        {
            "id": 3,
            "name": "根岸",
            "working_days": [5, 6],
            "closed_days": [2, 3],  # 定休日（水曜、木曜）
            "work_on_holidays": True,
            "min_staff": 1,
            "max_staff": 2,
            "part_time_priority": False,
            "flexible_staffing": False
        }
    ],
    "staff": [
        {"id": 1, "name": "田中太郎", "type": "社員", "max_days": 31, "assigned_locations": []},
        {"id": 2, "name": "鈴木花子", "type": "パート", "max_days": 12, "assigned_locations": [2]},
        {"id": 3, "name": "佐藤次郎", "type": "社員", "max_days": 31, "assigned_locations": []},
        {"id": 4, "name": "山田美咲", "type": "パート", "max_days": 10, "assigned_locations": [2]},
    ],
    "ng_days": {},
    "exceptions": {}
}


def get_firestore_client():
    """Firestoreクライアントを取得"""
    if not FIRESTORE_AVAILABLE:
        return None
    try:
        return firestore.Client()
    except Exception as e:
        print(f"Firestore接続エラー: {e}")
        return None


def ensure_data_dir():
    """データディレクトリを作成"""
    DATA_DIR.mkdir(exist_ok=True)


def load_data():
    """データを読み込む（Firestore優先、ローカルフォールバック）"""
    db = get_firestore_client()
    if db:
        try:
            doc = db.collection('settings').document('main').get()
            if doc.exists:
                return doc.to_dict()
        except Exception as e:
            print(f"Firestore読み込みエラー: {e}")

    # ローカルファイルにフォールバック
    ensure_data_dir()
    if DATA_FILE.exists():
        try:
            with open(DATA_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError):
            pass
    return deepcopy(DEFAULT_DATA)


def save_data(data):
    """データを保存（Firestoreとローカル両方）"""
    db = get_firestore_client()
    if db:
        try:
            db.collection('settings').document('main').set(data)
        except Exception as e:
            print(f"Firestore保存エラー: {e}")

    # ローカルにも保存（バックアップ）
    ensure_data_dir()
    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def get_locations():
    data = load_data()
    return data.get('locations', [])


def set_locations(locations):
    data = load_data()
    data['locations'] = locations
    save_data(data)


def get_staff():
    data = load_data()
    return data.get('staff', [])


def set_staff(staff):
    data = load_data()
    data['staff'] = staff
    save_data(data)


def get_ng_days():
    data = load_data()
    return data.get('ng_days', {})


def set_ng_days(ng_days):
    data = load_data()
    data['ng_days'] = ng_days
    save_data(data)


def get_exceptions():
    data = load_data()
    return data.get('exceptions', {})


def set_exceptions(exceptions):
    data = load_data()
    data['exceptions'] = exceptions
    save_data(data)


# =============================================================================
# シフト保存・管理機能（Firestore）
# =============================================================================

def save_shift_to_firestore(year, month, shift_data, staff_counts, ng_days_data, exceptions_data):
    """シフトをFirestoreに保存"""
    db = get_firestore_client()
    doc_id = f"{year}-{month:02d}"

    shift_doc = {
        "year": year,
        "month": month,
        "shift_data": shift_data,
        "staff_counts": staff_counts,
        "ng_days": ng_days_data,
        "exceptions": exceptions_data,
        "created_at": firestore.SERVER_TIMESTAMP if db else datetime.now().isoformat(),
        "updated_at": firestore.SERVER_TIMESTAMP if db else datetime.now().isoformat()
    }

    if db:
        try:
            db.collection('shifts').document(doc_id).set(shift_doc)
            return True
        except Exception as e:
            print(f"シフト保存エラー: {e}")
            return False

    # ローカルフォールバック
    shifts_file = DATA_DIR / 'shifts.json'
    ensure_data_dir()
    shifts = {}
    if shifts_file.exists():
        try:
            with open(shifts_file, 'r', encoding='utf-8') as f:
                shifts = json.load(f)
        except:
            pass

    shift_doc['created_at'] = datetime.now().isoformat()
    shift_doc['updated_at'] = datetime.now().isoformat()
    shifts[doc_id] = shift_doc

    with open(shifts_file, 'w', encoding='utf-8') as f:
        json.dump(shifts, f, ensure_ascii=False, indent=2)
    return True


def load_shift_from_firestore(year, month):
    """シフトをFirestoreから読み込み"""
    db = get_firestore_client()
    doc_id = f"{year}-{month:02d}"

    if db:
        try:
            doc = db.collection('shifts').document(doc_id).get()
            if doc.exists:
                data = doc.to_dict()
                # Firestore TimestampをISO形式に変換
                if data.get('created_at') and hasattr(data['created_at'], 'isoformat'):
                    data['created_at'] = data['created_at'].isoformat()
                if data.get('updated_at') and hasattr(data['updated_at'], 'isoformat'):
                    data['updated_at'] = data['updated_at'].isoformat()
                return data
        except Exception as e:
            print(f"シフト読み込みエラー: {e}")

    # ローカルフォールバック
    shifts_file = DATA_DIR / 'shifts.json'
    if shifts_file.exists():
        try:
            with open(shifts_file, 'r', encoding='utf-8') as f:
                shifts = json.load(f)
                return shifts.get(doc_id)
        except:
            pass
    return None


def delete_shift_from_firestore(year, month):
    """シフトをFirestoreから削除"""
    db = get_firestore_client()
    doc_id = f"{year}-{month:02d}"

    if db:
        try:
            db.collection('shifts').document(doc_id).delete()
            return True
        except Exception as e:
            print(f"シフト削除エラー: {e}")
            return False

    # ローカルフォールバック
    shifts_file = DATA_DIR / 'shifts.json'
    if shifts_file.exists():
        try:
            with open(shifts_file, 'r', encoding='utf-8') as f:
                shifts = json.load(f)
            if doc_id in shifts:
                del shifts[doc_id]
                with open(shifts_file, 'w', encoding='utf-8') as f:
                    json.dump(shifts, f, ensure_ascii=False, indent=2)
                return True
        except:
            pass
    return False


def list_saved_shifts():
    """保存済みシフト一覧を取得"""
    db = get_firestore_client()
    shifts_list = []

    if db:
        try:
            docs = db.collection('shifts').order_by('year', direction=firestore.Query.DESCENDING).order_by('month', direction=firestore.Query.DESCENDING).stream()
            for doc in docs:
                data = doc.to_dict()
                shifts_list.append({
                    "id": doc.id,
                    "year": data.get('year'),
                    "month": data.get('month'),
                    "updated_at": data.get('updated_at').isoformat() if hasattr(data.get('updated_at'), 'isoformat') else data.get('updated_at')
                })
            return shifts_list
        except Exception as e:
            print(f"シフト一覧取得エラー: {e}")

    # ローカルフォールバック
    shifts_file = DATA_DIR / 'shifts.json'
    if shifts_file.exists():
        try:
            with open(shifts_file, 'r', encoding='utf-8') as f:
                shifts = json.load(f)
            for doc_id, data in shifts.items():
                shifts_list.append({
                    "id": doc_id,
                    "year": data.get('year'),
                    "month": data.get('month'),
                    "updated_at": data.get('updated_at')
                })
            # 年月で降順ソート
            shifts_list.sort(key=lambda x: (x['year'], x['month']), reverse=True)
        except:
            pass
    return shifts_list

# =============================================================================
# ルーティング - ページ表示（認証必須）
# =============================================================================

@app.route('/locations')
@login_required
def locations_page():
    return render_template('locations.html', locations=get_locations())


@app.route('/staff')
@login_required
def staff_page():
    return render_template('staff.html', staff=get_staff(), locations=get_locations())


@app.route('/system')
@login_required
def system_page():
    return render_template('system.html')

# =============================================================================
# API - 拠点管理（認証必須）
# =============================================================================

@app.route('/api/locations', methods=['GET'])
@login_required
def api_get_locations():
    return jsonify(get_locations())

@app.route('/api/locations', methods=['POST'])
@login_required
def api_add_location():
    data = request.json
    locations = get_locations()

    max_id = max([loc['id'] for loc in locations], default=0)
    new_location = {
        "id": max_id + 1,
        "name": data.get('name', '新規拠点'),
        "working_days": data.get('working_days', [5, 6]),
        "closed_days": data.get('closed_days', []),  # 定休日
        "work_on_holidays": data.get('work_on_holidays', True),
        "min_staff": data.get('min_staff', 1),
        "max_staff": data.get('max_staff', 2),
        "part_time_priority": data.get('part_time_priority', False),
        "flexible_staffing": data.get('flexible_staffing', False)
    }
    locations.append(new_location)
    set_locations(locations)
    return jsonify(new_location)

@app.route('/api/locations/<int:loc_id>', methods=['PUT'])
@login_required
def api_update_location(loc_id):
    data = request.json
    locations = get_locations()

    for loc in locations:
        if loc['id'] == loc_id:
            loc['name'] = data.get('name', loc['name'])
            loc['working_days'] = data.get('working_days', loc.get('working_days', [5, 6]))
            loc['closed_days'] = data.get('closed_days', loc.get('closed_days', []))
            loc['work_on_holidays'] = data.get('work_on_holidays', loc.get('work_on_holidays', True))
            loc['min_staff'] = data.get('min_staff', loc.get('min_staff', 1))
            loc['max_staff'] = data.get('max_staff', loc.get('max_staff', 2))
            loc['part_time_priority'] = data.get('part_time_priority', loc['part_time_priority'])
            loc['flexible_staffing'] = data.get('flexible_staffing', loc['flexible_staffing'])
            set_locations(locations)
            return jsonify(loc)

    return jsonify({"error": "拠点が見つかりません"}), 404

@app.route('/api/locations/<int:loc_id>', methods=['DELETE'])
@login_required
def api_delete_location(loc_id):
    locations = get_locations()
    locations = [loc for loc in locations if loc['id'] != loc_id]
    set_locations(locations)
    return jsonify({"success": True})

# =============================================================================
# API - スタッフ管理（認証必須）
# =============================================================================

@app.route('/api/staff', methods=['GET'])
@login_required
def api_get_staff():
    return jsonify(get_staff())

@app.route('/api/staff', methods=['POST'])
@login_required
def api_add_staff():
    data = request.json
    staff_list = get_staff()

    max_id = max([s['id'] for s in staff_list], default=0)
    new_staff = {
        "id": max_id + 1,
        "name": data.get('name', '新規スタッフ'),
        "type": data.get('type', '社員'),
        "max_days": data.get('max_days', 31),
        "assigned_locations": data.get('assigned_locations', [])  # 所属拠点（パート用）
    }
    staff_list.append(new_staff)
    set_staff(staff_list)
    return jsonify(new_staff)

@app.route('/api/staff/<int:staff_id>', methods=['PUT'])
@login_required
def api_update_staff(staff_id):
    data = request.json
    staff_list = get_staff()

    for s in staff_list:
        if s['id'] == staff_id:
            s['name'] = data.get('name', s['name'])
            s['type'] = data.get('type', s['type'])
            s['max_days'] = data.get('max_days', s['max_days'])
            s['assigned_locations'] = data.get('assigned_locations', s.get('assigned_locations', []))
            set_staff(staff_list)
            return jsonify(s)

    return jsonify({"error": "スタッフが見つかりません"}), 404

@app.route('/api/staff/<int:staff_id>', methods=['DELETE'])
@login_required
def api_delete_staff(staff_id):
    staff_list = get_staff()
    staff_list = [s for s in staff_list if s['id'] != staff_id]
    set_staff(staff_list)
    return jsonify({"success": True})

# =============================================================================
# API - NG日・例外日管理（認証必須）
# =============================================================================

@app.route('/api/ng_days', methods=['GET'])
@login_required
def api_get_ng_days():
    return jsonify(get_ng_days())

@app.route('/api/ng_days', methods=['POST'])
@login_required
def api_set_ng_days():
    data = request.json
    set_ng_days(data)
    return jsonify({"success": True})

@app.route('/api/exceptions', methods=['GET'])
@login_required
def api_get_exceptions():
    return jsonify(get_exceptions())

@app.route('/api/exceptions/<int:year>/<int:month>', methods=['GET'])
@login_required
def api_get_month_exceptions(year, month):
    exceptions = get_exceptions()
    key = f"{year}-{month:02d}"
    return jsonify(exceptions.get(key, {}))

@app.route('/api/exceptions/<int:year>/<int:month>', methods=['POST'])
@login_required
def api_set_month_exceptions(year, month):
    data = request.json
    exceptions = get_exceptions()
    key = f"{year}-{month:02d}"
    exceptions[key] = data
    set_exceptions(exceptions)
    return jsonify({"success": True})

# =============================================================================
# API - 設定インポート/エクスポート・リセット（認証必須）
# =============================================================================

@app.route('/api/export_settings', methods=['GET'])
@login_required
def api_export_settings():
    data = load_data()
    data['export_date'] = datetime.now().isoformat()

    output = BytesIO()
    output.write(json.dumps(data, ensure_ascii=False, indent=2).encode('utf-8'))
    output.seek(0)

    return send_file(
        output,
        mimetype='application/json',
        as_attachment=True,
        download_name=f'shift_settings_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json'
    )

@app.route('/api/import_settings', methods=['POST'])
@login_required
def api_import_settings():
    if 'file' not in request.files:
        return jsonify({"error": "ファイルがありません"}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "ファイルが選択されていません"}), 400

    try:
        content = file.read().decode('utf-8')
        settings = json.loads(content)

        data = load_data()
        if 'locations' in settings:
            data['locations'] = settings['locations']
        if 'staff' in settings:
            data['staff'] = settings['staff']
        if 'ng_days' in settings:
            data['ng_days'] = settings['ng_days']
        if 'exceptions' in settings:
            data['exceptions'] = settings['exceptions']

        save_data(data)
        return jsonify({"success": True, "message": "設定をインポートしました"})
    except json.JSONDecodeError:
        return jsonify({"error": "無効なJSONファイルです"}), 400
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/reset', methods=['POST'])
@login_required
def api_reset():
    """設定をデフォルトにリセット"""
    save_data(deepcopy(DEFAULT_DATA))
    return jsonify({"success": True, "message": "設定をリセットしました"})


# =============================================================================
# API - シフト保存・管理（認証必須）
# =============================================================================

@app.route('/api/shifts', methods=['GET'])
@login_required
def api_list_shifts():
    """保存済みシフト一覧を取得"""
    shifts = list_saved_shifts()
    return jsonify(shifts)


@app.route('/api/shifts/<int:year>/<int:month>', methods=['GET'])
@login_required
def api_get_shift(year, month):
    """保存済みシフトを取得"""
    shift = load_shift_from_firestore(year, month)
    if shift:
        return jsonify(shift)
    return jsonify({"error": "シフトが見つかりません"}), 404


@app.route('/api/shifts/<int:year>/<int:month>', methods=['POST'])
@login_required
def api_save_shift(year, month):
    """シフトを保存"""
    data = request.json
    shift_data = data.get('shift_data', {})
    staff_counts = data.get('staff_counts', {})
    ng_days_data = data.get('ng_days', {})
    exceptions_data = data.get('exceptions', {})

    success = save_shift_to_firestore(year, month, shift_data, staff_counts, ng_days_data, exceptions_data)
    if success:
        return jsonify({"success": True, "message": f"{year}年{month}月のシフトを保存しました"})
    return jsonify({"error": "シフトの保存に失敗しました"}), 500


@app.route('/api/shifts/<int:year>/<int:month>', methods=['DELETE'])
@login_required
def api_delete_shift(year, month):
    """シフトを削除"""
    success = delete_shift_from_firestore(year, month)
    if success:
        return jsonify({"success": True, "message": f"{year}年{month}月のシフトを削除しました"})
    return jsonify({"error": "シフトの削除に失敗しました"}), 500

# =============================================================================
# カレンダー生成ヘルパー
# =============================================================================

def get_calendar_data(year, month, month_exceptions=None):
    locations = get_locations()
    if month_exceptions is None:
        exceptions = get_exceptions()
        key = f"{year}-{month:02d}"
        month_exceptions = exceptions.get(key, {})

    cal_data = []
    _, num_days = calendar.monthrange(year, month)

    for day in range(1, num_days + 1):
        d = date(year, month, day)
        weekday = d.weekday()
        is_holiday = jpholiday.is_holiday(d)
        holiday_name = jpholiday.is_holiday_name(d)
        date_str = d.isoformat()

        day_info = {
            "date": date_str,
            "day": day,
            "weekday": weekday,
            "weekday_name": ['月', '火', '水', '木', '金', '土', '日'][weekday],
            "is_holiday": is_holiday,
            "holiday_name": holiday_name,
            "locations": []
        }

        for loc in locations:
            loc_id_str = str(loc['id'])
            loc_exceptions = month_exceptions.get(loc_id_str, {"add": [], "remove": []})
            add_dates = loc_exceptions.get("add", [])
            remove_dates = loc_exceptions.get("remove", [])

            # 定休日チェック
            closed_days = loc.get('closed_days', [])
            is_closed_day = weekday in closed_days

            is_working = False
            if is_closed_day:
                is_working = False  # 定休日は営業しない
            elif is_holiday:
                is_working = loc.get('work_on_holidays', True)
            else:
                is_working = weekday in loc.get('working_days', [5, 6])

            if date_str in add_dates:
                is_working = True
            if date_str in remove_dates:
                is_working = False

            day_info['locations'].append({
                "id": loc['id'],
                "name": loc['name'],
                "is_working": is_working,
                "is_closed_day": is_closed_day,
                "min_staff": loc.get('min_staff', 1),
                "max_staff": loc.get('max_staff', 2),
                "part_time_priority": loc.get('part_time_priority', False),
                "flexible_staffing": loc.get('flexible_staffing', False)
            })

        cal_data.append(day_info)

    return cal_data

@app.route('/api/calendar/<int:year>/<int:month>', methods=['GET'])
@login_required
def api_get_calendar(year, month):
    cal_data = get_calendar_data(year, month)
    return jsonify(cal_data)

# =============================================================================
# シフト生成アルゴリズム
# =============================================================================

def generate_shift(year, month, ng_days_data, month_exceptions):
    locations = get_locations()
    staff_list = get_staff()
    cal_data = get_calendar_data(year, month, month_exceptions)

    staff_counts = {s['id']: 0 for s in staff_list}
    shift_result = {}

    for day_info in cal_data:
        date_str = day_info['date']
        shift_result[date_str] = {}

        for loc_info in day_info['locations']:
            loc_id = loc_info['id']
            shift_result[date_str][loc_id] = []

            if not loc_info['is_working']:
                continue

            min_required = loc_info['min_staff']
            max_allowed = loc_info['max_staff']
            assigned = []

            available_staff = []
            for s in staff_list:
                staff_id = s['id']
                staff_id_str = str(staff_id)

                # パートの場合、所属拠点チェック
                if s['type'] == 'パート':
                    assigned_locs = s.get('assigned_locations', [])
                    # 所属拠点が設定されていて、現在の拠点が含まれていなければスキップ
                    if assigned_locs and loc_id not in assigned_locs:
                        continue

                staff_ng_days = ng_days_data.get(staff_id_str, [])
                if date_str in staff_ng_days:
                    continue

                if staff_counts[staff_id] >= s['max_days']:
                    continue

                already_assigned_today = False
                for other_loc_id, other_assigned in shift_result[date_str].items():
                    if staff_id in other_assigned:
                        already_assigned_today = True
                        break
                if already_assigned_today:
                    continue

                available_staff.append(s)

            if loc_info['part_time_priority']:
                part_timers = [s for s in available_staff if s['type'] == 'パート']
                regulars = [s for s in available_staff if s['type'] == '社員']

                part_timers.sort(key=lambda s: staff_counts[s['id']])
                regulars.sort(key=lambda s: staff_counts[s['id']])

                # パート優先枠でも、まず社員を1人確保する（パートのみにはしない）
                for s in regulars:
                    if len(assigned) >= 1:
                        break
                    assigned.append(s['id'])
                    staff_counts[s['id']] += 1

                # パートは最大1人まで（2人は入らない）
                for s in part_timers:
                    if len(assigned) >= max_allowed:
                        break
                    assigned.append(s['id'])
                    staff_counts[s['id']] += 1
                    break  # パートは1人追加したら終了

                # flexible_staffing: パートがいなければ最小人数を1に減らす
                if loc_info['flexible_staffing']:
                    part_count = len([sid for sid in assigned
                                    if any(s['id'] == sid and s['type'] == 'パート' for s in staff_list)])
                    if part_count == 0 and min_required > 1:
                        min_required = 1

                # 残りの枠を社員で埋める（min_requiredまで）
                for s in regulars:
                    if len(assigned) >= max_allowed:
                        break
                    if len(assigned) >= min_required:
                        break
                    if s['id'] in assigned:
                        continue
                    assigned.append(s['id'])
                    staff_counts[s['id']] += 1
            else:
                available_staff.sort(key=lambda s: staff_counts[s['id']])

                for s in available_staff:
                    if len(assigned) >= max_allowed:
                        break
                    assigned.append(s['id'])
                    staff_counts[s['id']] += 1

            shift_result[date_str][loc_id] = assigned

    return {
        "year": year,
        "month": month,
        "shift": shift_result,
        "staff_counts": staff_counts
    }

@app.route('/api/generate_shift', methods=['POST'])
@login_required
def api_generate_shift():
    data = request.json
    year = data.get('year', datetime.now().year)
    month = data.get('month', datetime.now().month)
    ng_days_data = data.get('ng_days', get_ng_days())
    month_exceptions = data.get('exceptions', {})

    exceptions = get_exceptions()
    key = f"{year}-{month:02d}"
    exceptions[key] = month_exceptions
    set_exceptions(exceptions)

    result = generate_shift(year, month, ng_days_data, month_exceptions)
    return jsonify(result)

# =============================================================================
# Excel出力（シンプルなカレンダー形式）
# =============================================================================

def create_excel_shift(year, month, shift_data, month_exceptions):
    locations = get_locations()
    staff_list = get_staff()
    staff_dict = {s['id']: s['name'] for s in staff_list}
    num_locations = len(locations)

    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}年{month}月シフト表"

    # スタイル定義
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    saturday_fill = PatternFill(start_color='DEEAF6', end_color='DEEAF6', fill_type='solid')
    sunday_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
    holiday_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
    closed_day_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
    loc_name_fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')

    saturday_font = Font(color='0000FF', bold=True)
    sunday_font = Font(color='FF0000', bold=True)

    cal_data = get_calendar_data(year, month, month_exceptions)
    weekday_headers = ['日', '月', '火', '水', '木', '金', '土']

    # カレンダーデータを週単位でグループ化
    weeks = []
    current_week = [None] * 7
    for day_info in cal_data:
        weekday_jp = (day_info['weekday'] + 1) % 7
        current_week[weekday_jp] = day_info
        if weekday_jp == 6:
            weeks.append(current_week)
            current_week = [None] * 7
    if any(d is not None for d in current_week):
        weeks.append(current_week)

    # 列幅設定
    ws.column_dimensions['A'].width = 6
    for col in range(2, 9):
        ws.column_dimensions[get_column_letter(col)].width = 12

    # ヘッダー行: 空白 + 曜日
    ws.cell(row=1, column=1, value="").border = thin_border
    for col, day_name in enumerate(weekday_headers, start=2):
        cell = ws.cell(row=1, column=col, value=day_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if col == 2:
            cell.fill = PatternFill(start_color='dc3545', end_color='dc3545', fill_type='solid')
        elif col == 8:
            cell.fill = PatternFill(start_color='0d6efd', end_color='0d6efd', fill_type='solid')
    ws.row_dimensions[1].height = 20

    current_row = 2

    for week_idx, week in enumerate(weeks):
        # 全拠点が定休日かチェック
        all_closed_days = []
        for day_info in week:
            if day_info:
                all_closed = all(loc_info.get('is_closed_day', False) for loc_info in day_info['locations'])
                all_closed_days.append(all_closed)
            else:
                all_closed_days.append(False)

        # 日付行
        date_row = current_row
        month_cell = ws.cell(row=date_row, column=1)
        if week_idx == 0:
            month_cell.value = f"{month}月"
            month_cell.font = Font(size=12, bold=True)
        else:
            month_cell.value = ""
        month_cell.alignment = Alignment(horizontal='center', vertical='center')
        month_cell.border = thin_border

        for col_idx, day_info in enumerate(week):
            col = col_idx + 2
            cell = ws.cell(row=date_row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            if day_info:
                day = day_info['day']
                weekday_jp = (day_info['weekday'] + 1) % 7
                is_holiday = day_info['is_holiday']
                holiday_name = day_info['holiday_name']

                if is_holiday and holiday_name:
                    cell.value = f"{day}\n{holiday_name}"
                else:
                    cell.value = day

                if is_holiday or weekday_jp == 0:
                    cell.font = sunday_font
                    cell.fill = sunday_fill if not is_holiday else holiday_fill
                elif weekday_jp == 6:
                    cell.font = saturday_font
                    cell.fill = saturday_fill
                else:
                    cell.font = Font(bold=True)

        ws.row_dimensions[date_row].height = 30

        # 各拠点行
        for loc_idx, loc in enumerate(locations):
            loc_row = date_row + 1 + loc_idx

            loc_name_cell = ws.cell(row=loc_row, column=1, value=loc['name'])
            loc_name_cell.font = Font(size=9, bold=True)
            loc_name_cell.alignment = Alignment(horizontal='center', vertical='center')
            loc_name_cell.border = thin_border
            loc_name_cell.fill = loc_name_fill

            for col_idx, day_info in enumerate(week):
                col = col_idx + 2
                cell = ws.cell(row=loc_row, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                if day_info:
                    date_str = day_info['date']
                    loc_info = next((l for l in day_info['locations'] if l['id'] == loc['id']), None)

                    # 背景色は名前部分には付けない（土日祝でも色なし）

                    # 定休日
                    if all_closed_days[col_idx]:
                        if loc_idx == 0:
                            if num_locations > 1:
                                ws.merge_cells(start_row=loc_row, start_column=col,
                                             end_row=loc_row + num_locations - 1, end_column=col)
                            cell.value = "定休日"
                            cell.font = Font(size=11, bold=True)
                            cell.fill = closed_day_fill
                    elif loc_info and loc_info.get('is_closed_day', False):
                        cell.value = "休"
                        cell.fill = closed_day_fill
                    elif loc_info and not loc_info['is_working']:
                        cell.value = ""
                    else:
                        shift_for_day = shift_data.get(date_str, {})
                        assigned_ids = shift_for_day.get(str(loc['id']), [])
                        if assigned_ids:
                            names = [staff_dict.get(int(sid) if isinstance(sid, str) else sid, '?')
                                   for sid in assigned_ids if sid]
                            cell.value = "\n".join(names) if names else "-"
                            cell.font = Font(size=9, bold=True)  # 名前は太字
                        else:
                            cell.value = "-"

            ws.row_dimensions[loc_row].height = 26

        current_row = date_row + 1 + num_locations

    return wb

@app.route('/api/export_excel', methods=['POST'])
@login_required
def api_export_excel():
    data = request.json
    year = data.get('year', datetime.now().year)
    month = data.get('month', datetime.now().month)
    month_exceptions = data.get('exceptions', {})

    # 編集済みシフトデータがあればそれを使用、なければ生成
    shift_data = data.get('shift_data')
    if not shift_data:
        ng_days_data = data.get('ng_days', get_ng_days())
        result = generate_shift(year, month, ng_days_data, month_exceptions)
        shift_data = result['shift']

    wb = create_excel_shift(year, month, shift_data, month_exceptions)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"shift_{year}_{month:02d}.xlsx"

    response = send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
    # ダウンロード警告を軽減するためのヘッダー
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['Cache-Control'] = 'no-cache'
    return response

# =============================================================================
# PDF出力
# =============================================================================

def get_japanese_font():
    """日本語フォントを取得"""
    # Windowsの場合
    font_paths = [
        'C:/Windows/Fonts/msgothic.ttc',  # MS ゴシック
        'C:/Windows/Fonts/meiryo.ttc',     # メイリオ
        'C:/Windows/Fonts/YuGothM.ttc',    # 游ゴシック
    ]
    for font_path in font_paths:
        if os.path.exists(font_path):
            return font_path
    return None

def create_pdf_shift(year, month, shift_data, month_exceptions):
    """PDF形式のシフト表を作成（A4横1枚に収める - Canvas直接描画）"""
    from reportlab.pdfgen import canvas

    locations = get_locations()
    staff_list = get_staff()
    staff_dict = {s['id']: s['name'] for s in staff_list}
    cal_data = get_calendar_data(year, month, month_exceptions)
    num_locations = len(locations)

    output = BytesIO()

    # A4横: 297mm x 210mm (841.89 x 595.28 points)
    page_width, page_height = landscape(A4)

    c = canvas.Canvas(output, pagesize=landscape(A4))

    # 日本語フォント登録（通常と太字）
    font_path = get_japanese_font()
    font_name = 'Helvetica'
    font_name_bold = 'Helvetica-Bold'
    if font_path:
        try:
            pdfmetrics.registerFont(TTFont('JapaneseFont', font_path))
            font_name = 'JapaneseFont'
            # 太字フォント（メイリオBoldまたは游ゴシックBold）
            bold_paths = [
                'C:/Windows/Fonts/meiryob.ttc',   # メイリオBold
                'C:/Windows/Fonts/YuGothB.ttc',   # 游ゴシックBold
                'C:/Windows/Fonts/msgothic.ttc',  # MS ゴシック（太字なし、同じフォント使用）
            ]
            font_name_bold = font_name  # デフォルトは通常フォント
            for bold_path in bold_paths:
                if os.path.exists(bold_path):
                    try:
                        pdfmetrics.registerFont(TTFont('JapaneseFontBold', bold_path))
                        font_name_bold = 'JapaneseFontBold'
                        break
                    except:
                        pass
        except:
            font_name = 'Helvetica'
            font_name_bold = 'Helvetica-Bold'

    # カレンダーデータを週単位でグループ化
    weeks = []
    current_week = [None] * 7
    for day_info in cal_data:
        weekday_jp = (day_info['weekday'] + 1) % 7
        current_week[weekday_jp] = day_info
        if weekday_jp == 6:
            weeks.append(current_week)
            current_week = [None] * 7
    if any(d is not None for d in current_week):
        weeks.append(current_week)

    # レイアウト計算
    margin = 10  # points
    table_x = margin
    table_y = margin
    table_width = page_width - 2 * margin
    table_height = page_height - 2 * margin

    # 列幅
    first_col_width = 40
    other_col_width = (table_width - first_col_width) / 7
    col_widths = [first_col_width] + [other_col_width] * 7

    # 行数計算: ヘッダー1行 + 週数×(日付行1 + 拠点数)
    num_rows = 1 + len(weeks) * (1 + num_locations)
    row_height = table_height / num_rows

    # フォントサイズ（大きめに設定）
    font_size = min(10, row_height * 0.7)

    weekday_headers = ['', '日', '月', '火', '水', '木', '金', '土']

    def get_col_x(col_idx):
        x = table_x
        for i in range(col_idx):
            x += col_widths[i]
        return x

    def draw_cell(row_idx, col_idx, text, bg_color=None, text_color=None, bold=False, row_span=1):
        x = get_col_x(col_idx)
        y = page_height - table_y - (row_idx + row_span) * row_height
        w = col_widths[col_idx]
        h = row_height * row_span  # 複数行にまたがる場合

        # 背景色
        if bg_color:
            c.setFillColor(bg_color)
            c.rect(x, y, w, h, fill=1, stroke=0)

        # 枠線
        c.setStrokeColor(colors.black)
        c.setLineWidth(0.5)
        c.rect(x, y, w, h, fill=0, stroke=1)

        # テキスト
        if text:
            # 太字指定の場合は太字フォントを使用
            use_font = font_name_bold if bold else font_name
            c.setFont(use_font, font_size)
            if text_color:
                c.setFillColor(text_color)
            else:
                c.setFillColor(colors.black)
            # テキスト中央配置
            text_width = c.stringWidth(str(text), use_font, font_size)
            text_x = x + (w - text_width) / 2
            text_y = y + (h - font_size) / 2 + font_size * 0.2
            c.drawString(text_x, text_y, str(text))

    # ヘッダー行描画
    for col_idx, header in enumerate(weekday_headers):
        if col_idx == 0:
            bg = colors.HexColor('#4472C4')
        elif col_idx == 1:  # 日曜
            bg = colors.HexColor('#dc3545')
        elif col_idx == 7:  # 土曜
            bg = colors.HexColor('#0d6efd')
        else:
            bg = colors.HexColor('#4472C4')
        draw_cell(0, col_idx, header, bg_color=bg, text_color=colors.white)

    # データ行描画
    current_row = 1
    for week_idx, week in enumerate(weeks):
        # 全拠点が定休日かチェック
        all_closed_days = []
        for day_info in week:
            if day_info:
                all_closed = all(loc_info.get('is_closed_day', False) for loc_info in day_info['locations'])
                all_closed_days.append(all_closed)
            else:
                all_closed_days.append(False)

        # 日付行
        # 左端セル（月名）
        month_text = f"{month}月" if week_idx == 0 else ""
        draw_cell(current_row, 0, month_text, bg_color=colors.HexColor('#f8f9fa'))

        for col_idx, day_info in enumerate(week):
            col = col_idx + 1
            if day_info:
                day = day_info['day']
                weekday_jp = (day_info['weekday'] + 1) % 7
                is_holiday = day_info.get('is_holiday', False)
                holiday_name = day_info.get('holiday_name', '')

                cell_text = str(day)
                if is_holiday and holiday_name:
                    cell_text = f"{day} {holiday_name}"

                if is_holiday or weekday_jp == 0:
                    bg = colors.HexColor('#ffe6e6')
                    txt_color = colors.HexColor('#dc3545')
                elif weekday_jp == 6:
                    bg = colors.HexColor('#e6f0ff')
                    txt_color = colors.HexColor('#0d6efd')
                else:
                    bg = None
                    txt_color = colors.black
                draw_cell(current_row, col, cell_text, bg_color=bg, text_color=txt_color)
            else:
                draw_cell(current_row, col, "")
        current_row += 1

        # 定休日の列を先に結合セルとして描画
        first_loc_row = current_row
        for col_idx, day_info in enumerate(week):
            col = col_idx + 1
            if day_info and all_closed_days[col_idx]:
                # 全拠点が定休日の場合、結合セルを描画
                draw_cell(first_loc_row, col, "定休日",
                         bg_color=colors.HexColor('#f0f0f0'),
                         row_span=num_locations)

        # 各拠点行
        for loc_idx, loc in enumerate(locations):
            # 拠点名
            draw_cell(current_row, 0, loc['name'], bg_color=colors.HexColor('#f8f9fa'))

            for col_idx, day_info in enumerate(week):
                col = col_idx + 1
                # 全拠点定休日の列はスキップ（既に結合セルで描画済み）
                if all_closed_days[col_idx]:
                    continue

                if day_info:
                    loc_info = next((l for l in day_info['locations'] if l['id'] == loc['id']), None)

                    # 背景色は基本なし（土日祝でも色なし）
                    bg = None

                    # セル内容
                    is_name = False  # 名前かどうかのフラグ
                    if loc_info and loc_info.get('is_closed_day', False):
                        # 一部拠点のみ定休日
                        cell_text = "休"
                        bg = colors.HexColor('#f0f0f0')
                    elif loc_info and not loc_info['is_working']:
                        cell_text = "休"
                    else:
                        assigned_ids = shift_data.get(day_info['date'], {}).get(str(loc['id']), [])
                        if assigned_ids:
                            names = [staff_dict.get(int(sid) if isinstance(sid, str) else sid, '?')
                                   for sid in assigned_ids if sid]
                            cell_text = "/".join(names) if names else "-"
                            is_name = bool(names)  # 名前がある場合は太字
                        else:
                            cell_text = "-"

                    draw_cell(current_row, col, cell_text, bg_color=bg, bold=is_name)
                else:
                    draw_cell(current_row, col, "")
            current_row += 1

    c.save()
    output.seek(0)
    return output

@app.route('/api/export_pdf', methods=['POST'])
@login_required
def api_export_pdf():
    data = request.json
    year = data.get('year', datetime.now().year)
    month = data.get('month', datetime.now().month)
    month_exceptions = data.get('exceptions', {})

    # 編集済みシフトデータがあればそれを使用
    shift_data = data.get('shift_data')
    if not shift_data:
        ng_days_data = data.get('ng_days', get_ng_days())
        result = generate_shift(year, month, ng_days_data, month_exceptions)
        shift_data = result['shift']

    output = create_pdf_shift(year, month, shift_data, month_exceptions)

    filename = f"shift_{year}_{month:02d}.pdf"

    response = send_file(
        output,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['Cache-Control'] = 'no-cache'
    return response

# =============================================================================
# メイン
# =============================================================================

if __name__ == '__main__':
    # 初回起動時にデータファイルがなければ作成
    if not DATA_FILE.exists():
        save_data(deepcopy(DEFAULT_DATA))
        print(f"データファイルを作成しました: {DATA_FILE}")

    app.run(debug=True, host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))
